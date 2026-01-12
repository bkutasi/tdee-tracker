#!/usr/bin/env python3
"""
Extract TDEE data from Excel file and convert to app import format.
Excel structure:
- Column B: Date (Saturday = end of week)
- Column C: Type ('Weight' or 'Cal.')
- Columns D-J: Mon-Sun values for that week (D=Mon, J=Sun)
"""

import openpyxl
import json
import datetime
import os
import sys

def extract():
    # Try multiple locations for the Excel file
    possible_paths = [
        'Improved_TDEE_Tracker.xlsx',
        '/home/bkutasi/Downloads/Improved_TDEE_Tracker.xlsx',
        '../Improved_TDEE_Tracker.xlsx'
    ]
    
    input_file = None
    for path in possible_paths:
        if os.path.exists(path):
            input_file = path
            break
    
    if not input_file:
        print(f"Error: Could not find Improved_TDEE_Tracker.xlsx")
        print(f"Tried: {possible_paths}")
        return
    
    output_file = 'import_data.json'
    
    print(f"Reading from: {input_file}")
    wb = openpyxl.load_workbook(input_file, data_only=True)
    sheet = wb['TDEE']
    
    import_entries = {}
    
    # Iterate looking for Date/Weight rows
    # The structure pattern is:
    # Row N-1: Calories
    # Row N: Date (Col B), Weight (Col C='Weight')
    # Week is Sunday (Col D) to Saturday (Col J/Date B)
    
    for row in range(12, 100):
        cell_date = sheet.cell(row=row, column=2).value
        cell_type = sheet.cell(row=row, column=3).value
        
        # We process key on the Weight row (which has the date)
        if isinstance(cell_date, datetime.datetime) and cell_type == 'Weight':
            saturday_date = cell_date.date()
            
            # Weight values in CURRENT row (D-J)
            w_vals = [sheet.cell(row=row, column=c).value for c in range(4, 11)]
            
            # Calorie values in PREVIOUS row (D-J)
            c_vals = [sheet.cell(row=row-1, column=c).value for c in range(4, 11)]
            
            # D=Sunday, J=Saturday (Date B)
            # D = Saturday - 6 days
            
            for i in range(7):
                # i=0 is Dim (Sunday), i=6 is Sat
                days_before_sat = 6 - i
                actual_date = saturday_date - datetime.timedelta(days=days_before_sat)
                date_str = actual_date.isoformat()
                
                w = w_vals[i] 
                c = c_vals[i]
                
                def safe_float(val):
                    if val is None:
                        return None
                    try:
                        return float(val)
                    except (ValueError, TypeError):
                        return None
                
                w_float = safe_float(w)
                c_float = safe_float(c)
                
                # Only add if there is data
                if w_float is not None or c_float is not None:
                    import_entries[date_str] = {
                        "weight": w_float,
                        "calories": c_float,
                        "notes": "Imported from Excel",
                        "updatedAt": datetime.datetime.now(datetime.timezone.utc).isoformat()
                    }

    # Sort entries by date
    sorted_entries = dict(sorted(import_entries.items()))
    
    # Format for Storage.importData
    data_to_export = {
        "version": 1,
        "exportedAt": datetime.datetime.now(datetime.timezone.utc).isoformat(),
        "settings": {
            "weightUnit": "kg",
            "calorieUnit": "cal"
        },
        "entries": sorted_entries
    }
    
    with open(output_file, 'w') as f:
        json.dump(data_to_export, f, indent=2)
    
    # Print summary
    dates = sorted(import_entries.keys())
    if dates:
        print(f"Successfully exported {len(import_entries)} entries to {output_file}")
        print(f"Date range: {dates[0]} to {dates[-1]}")
    else:
        print("No entries found!")

if __name__ == "__main__":
    extract()
